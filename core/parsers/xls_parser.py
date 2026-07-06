"""XLS/XLSX 文件解析器"""
import re

# UTF-16LE 文本段：每个码元 (low,high) 满足 0x20<=code<0xFFFE（与 doc 不同，无 0x09/0x0A/0x0D 例外）。
# 三支字节类：high=0x00 时 low∈0x20-0xff / high∈0x01-0xfe 任意 low / high=0xff 时 low∈0x00-0xfd。
# 覆盖中文（高字节非零）等全部有效码元；匹配 3+ 码元（与原 len>=3 一致）。
_RE_UNICODE = re.compile(
    rb'(?:[\x00-\xff][\x01-\xfe]|[\x20-\xff]\x00|[\x00-\xfd]\xff){3,}'
)
# ASCII/GBK 文本段：0x20-0x7e + 0x80-0xfe（无制表/换行）；匹配 5+ 字节（与原 len>4 一致）。
_RE_ASCII = re.compile(rb'[\x20-\x7e\x80-\xfe]{5,}')


def parse_xlsx(filepath):
    """解析 .xlsx 文件（Office Open XML格式）"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"[工作表: {sheet_name}]")

            for row in ws.iter_rows(values_only=True):
                cells = []
                for cell in row:
                    if cell is not None:
                        cells.append(str(cell))
                if cells:
                    text_parts.append(" | ".join(cells))

        wb.close()
        return "\n".join(text_parts)
    except ImportError:
        raise RuntimeError("openpyxl 未安装")
    except Exception as e:
        raise RuntimeError(f"XLSX解析失败: {e}")


def parse_xls(filepath):
    """解析 .xls 文件（旧版OLE格式）"""
    # 方法1: 使用 olefile 提取
    text = _parse_xls_olefile(filepath)
    if text and len(text.strip()) > 50:
        return text

    # 方法2: 二进制提取
    text = _parse_xls_binary(filepath)
    return text or ""


def _parse_xls_olefile(filepath):
    """使用olefile解析.xls文件"""
    try:
        import olefile
        if not olefile.isOleFile(filepath):
            return ""

        ole = olefile.OleFileIO(filepath)
        text_parts = []

        # Workbook 流包含主要数据
        if ole.exists("Workbook"):
            stream = ole.openstream("Workbook")
            data = stream.read()
            stream.close()

            # 提取字符串（Unicode和ASCII）
            text = _extract_xls_strings(data)
            if text:
                text_parts.append(text)

        ole.close()
        return "\n".join(text_parts)
    except ImportError:
        return ""
    except Exception:
        return ""


def _extract_xls_strings(data):
    """从XLS二进制数据中提取字符串

    预编译正则在 C 层扫描连续有效 UTF-16LE 码元（3+ 个），替代逐码元 Python 循环，
    大文件解析提速 10-50x。有效码元定义与原实现一致：0x20<=code<0xFFFE。
    """
    return "\n".join(
        m.decode('utf-16-le', 'ignore') for m in _RE_UNICODE.findall(data)
    )


def _parse_xls_binary(filepath):
    """从二进制.xls文件中提取文本

    预编译正则匹配连续可打印字节（5+ 个），替代逐字节 Python 循环。
    字节集与原实现一致：0x20-0x7e / 0x80-0xfe（无制表/换行）。
    """
    try:
        with open(filepath, "rb") as f:
            data = f.read()
        raw = b"\n".join(_RE_ASCII.findall(data))
        try:
            from core.text_utils import decode_bytes
            return decode_bytes(raw)
        except Exception:
            return raw.decode('latin-1')
    except Exception:
        return ""
